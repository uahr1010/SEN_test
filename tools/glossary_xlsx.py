# -*- coding: utf-8 -*-
"""용어집 JSON ↔ 엑셀 변환기.

원본은 어디까지나 content/glossary.json 입니다.
엑셀은 여러 사람이 한꺼번에 검토하거나 인쇄할 때 쓰는 사본입니다.

    # JSON → 엑셀 (검토용으로 뽑기)
    python tools/glossary_xlsx.py export

    # 엑셀 → JSON (엑셀에서 손본 내용을 되돌려 넣기)
    python tools/glossary_xlsx.py import

`import` 는 glossary.json 을 통째로 덮어씁니다.
Pages CMS 에서 편집한 내용이 있다면 먼저 export 로 최신본을 받아 두세요.
"""
import io
import json
import os
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
JSON_PATH = os.path.join(ROOT, 'content', 'glossary.json')
XLSX_PATH = os.path.join(HERE, 'glossary.xlsx')

COLS = [('ko', '한국어'), ('en', 'English'), ('ja', '日本語'), ('zh', '中文')]


def export():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    doc = json.load(io.open(JSON_PATH, encoding='utf-8'))
    terms = doc.get('terms', [])

    wb = Workbook()
    ws = wb.active
    ws.title = '용어집'
    ws.append([label for _, label in COLS])

    for c in range(1, len(COLS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical='center')

    for t in terms:
        ws.append([t.get(key, '') for key, _ in COLS])

    for i, (_, _) in enumerate(COLS, start=1):
        ws.column_dimensions[chr(64 + i)].width = 46
    ws.freeze_panes = 'A2'

    wb.save(XLSX_PATH)
    print('export: %d개 용어 → %s' % (len(terms), XLSX_PATH))


def import_():
    import openpyxl

    if not os.path.exists(XLSX_PATH):
        sys.exit('엑셀이 없습니다: %s\n먼저 export 로 만들어 주세요.' % XLSX_PATH)

    ws = openpyxl.load_workbook(XLSX_PATH, data_only=True).worksheets[0]
    rows = list(ws.iter_rows(min_row=2, max_col=len(COLS), values_only=True))

    terms, skipped = [], 0
    for r in rows:
        entry = {}
        for i, (key, _) in enumerate(COLS):
            v = r[i] if i < len(r) else None
            entry[key] = str(v).strip() if v is not None else ''
        if not entry['ko']:
            skipped += 1
            continue
        terms.append(entry)

    empty = [t['ko'] for t in terms if not all(t[k] for k, _ in COLS)]

    doc = json.load(io.open(JSON_PATH, encoding='utf-8'))
    doc['terms'] = terms
    io.open(JSON_PATH, 'w', encoding='utf-8').write(
        json.dumps(doc, ensure_ascii=False, indent=2) + u'\n')

    print('import: %d개 용어 → %s' % (len(terms), JSON_PATH))
    if skipped:
        print('  한국어가 비어 건너뛴 행: %d' % skipped)
    if empty:
        print('  ⚠ 번역이 덜 채워진 용어 %d개: %s' % (len(empty), ', '.join(empty[:5])))


if __name__ == '__main__':
    cmd = sys.argv[1] if len(sys.argv) > 1 else ''
    if cmd == 'export':
        export()
    elif cmd == 'import':
        import_()
    else:
        sys.exit(__doc__)
